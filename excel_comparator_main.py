import openpyxl
import copy

OUTPUT_FILE = "results.xlsx"
OUTPUT_WRITE_HEADER = True
OUTPUT_WRITE_DOUBLON = True
OUTPUT_WRITE_NO_SERIAL = True
OUTPUT_WRITE_SHORT_SERIAL = True

INVENTAIRE_PREFIX = "inv"
INVENTAIRE_FILENAME = "inventaire.xlsx"
INVENTAIRE_HEADER_LINE_COUNT = 1
INVENTAIRE_END_LINE = 909
INVENTAIRE_SERIAL_COLUMN = 'P'
INVENTAIRE_OUTPUT_COLUMNS = [
{"label": "serial",  "col":'P'}, 
{"label":  "constructeur",  "col": 'B'}, 
{"label":  "imo",  "col":  'O'}, 
{"label": "status",  "col":'Q'}
]

PARC_PREFIXE = "parc"
PARC_FILENAME  = "parc.xlsx"
PARC_HEADER_LINE_COUNT = 4
PARC_END_LINE = 810
PARC_SERIAL_COLUMN = 'Q'
PARC_OUTPUT_COLUMNS = [
{"label": "serial",  "col":'Q'}, 
{"label": "constructeur",  "col": 'J'}, 
{"label": "ref_constructeur",  "col": 'K'}, 
{"label": "designation",  "col": 'L'}
]

SERIALIZER_UNEXPECTED_CHAR = {' ',  '^',  '"',  '$',  '*',  '%',  '-', '−',  '\''} #etc
SERIALIZER_AREA_REMOVAL_BEGIN_CHAR = {'(',  '[',  '{'}
SERIALIZER_AREA_REMOVAL_END_CHAR = {')', ']', '}'}
SERIALIZER_NO_SERIAL_STRING = {"−",  "N/A"}

SERIAL_MIN_LENGTH = 5

def serialize_serial(serial):
    erase_area = False
    enter_char_index = 0
    
    test = copy.deepcopy(serial)
    
    serial = list(str(serial))
    index = 0
    
    while index < len(serial):
        char = serial[index]
        if char in SERIALIZER_UNEXPECTED_CHAR and erase_area == False:
            del serial[index]
            index -= 1
        elif (char in SERIALIZER_AREA_REMOVAL_BEGIN_CHAR) and erase_area == False:
            enter_char_index = index
            erase_area = True
        elif char in SERIALIZER_AREA_REMOVAL_END_CHAR:
            if erase_area == True:
                del serial[enter_char_index:index]
                erase_area = False
                index -= (index - enter_char_index + 1)
            else:
                del serial[index]
                index -=1
        index += 1
    proper_serial =  "".join(serial)
    if proper_serial in SERIALIZER_NO_SERIAL_STRING:
        return ""
    else:
        if test != proper_serial:
            print(test,  " - ",  proper_serial)
        return proper_serial
    
def serialize(sheet,  begin,  end, column,  serial,  doublon,  none,  short,  min_length):
    for i in range(1 + begin,  end + 1):
        current_cell = sheet[column + str(i)]
        if current_cell.value != None:
            proper_serial = serialize_serial(current_cell.value)
            if proper_serial == "":
                none.append(i)
            elif proper_serial not in serial.keys():
                if len(proper_serial) < min_length:
                    short[proper_serial] = i
                else:
                    serial[proper_serial] = i
            else:
                try:
                    doublon[proper_serial].append(i)
                except:
                    doublon[proper_serial] = []
                    doublon[proper_serial].append(i)
                if serial[proper_serial] not in doublon[proper_serial]:
                    doublon[proper_serial].append(serial[proper_serial])
        else:
            none.append(i)

# Load Excel files
inventaire_workbook = openpyxl.load_workbook(INVENTAIRE_FILENAME)
inventaire_sheet = inventaire_workbook.active
inventaire_key_dict = {}
inventaire_key_dict_doublon = {}
inventaire_key_dict_short = {}
inventaire_no_key_dict = []
inventaire_no_match = {}

parc_workbook = openpyxl.load_workbook(PARC_FILENAME)
parc_sheet = parc_workbook.active
parc_key_dict = {}
parc_key_dict_doublon = {}
parc_key_dict_short = {}
parc_no_key_dict = []
parc_no_match = {}

# Serialize key:
serialize(inventaire_sheet, INVENTAIRE_HEADER_LINE_COUNT,  INVENTAIRE_END_LINE,  INVENTAIRE_SERIAL_COLUMN,  inventaire_key_dict,  inventaire_key_dict_doublon,  inventaire_no_key_dict,  inventaire_key_dict_short,  SERIAL_MIN_LENGTH)
serialize(parc_sheet,  PARC_HEADER_LINE_COUNT,  PARC_END_LINE,  PARC_SERIAL_COLUMN,  parc_key_dict,  parc_key_dict_doublon,  parc_no_key_dict,  parc_key_dict_short,  SERIAL_MIN_LENGTH)

results = [] # Key -> inv_row; Value -> parc_row
for inv_key in inventaire_key_dict.keys():
    hasMatched = False
    for parc_key in parc_key_dict.keys():
        if inv_key in parc_key or parc_key in inv_key:
            results.append({
                "inv_key": inv_key, 
                "parc_key": parc_key, 
                "inv_row": inventaire_key_dict[inv_key], 
                "parc_row": parc_key_dict[parc_key]
            })
            hasMatched = True
    if not hasMatched:
        inventaire_no_match[inv_key] = inventaire_key_dict[inv_key]
for parc_key in parc_key_dict.keys():
    hasMatched = False
    for result in results:
        if parc_key in result.values():
            hasMatched = True
    if not hasMatched:
        parc_no_match[parc_key] = parc_key_dict[parc_key]


class P_Struct:
    def __init__(self):
        self.data = []
        self.header = None
        
    def add_empty_rows(self,  until):
        for i in range(until - len(self.data) + 1):
                self.data.append([])
                
        
    def write_header(self,  header,  mode='a',  key=None):
        if mode == 'a':
            try:
                self.data[0]
            except:
                self.data.append([])
        elif mode == 'w':
            self.data[0] = []
            
        self.header = self.data[0]
        if key is not None:
            for label in header:
                    self.header.append(label[key])
        else:
            for label in header:
                    self.header.append(label)
                
    def add_data_row(self,  row):
        self.data.append([])
        current_row = self.data[-1]
        for cell in row:
            current_row.append(cell)
            
    def write_data_row(self,  row,  at,  mode='a'):
        if len(self.data) - 1 < at:
            # Il faut d'abord créer la rangée
            self.add_empty_rows(at)
            
        current_row = self.data[at]
        if mode == 'w':
            current_row = []
            
        for cell in row:
            current_row.append(cell)
  
    def write_in_excel_sheet(self,  output_sheet):
        for row in range(0,  len(self.data)):
            for col in range(0,  len(self.data[row])):
                output_sheet.cell(row=row + 1,  column=col + 1,  value=self.data[row][col])
                
        return output_sheet

output_workbook = openpyxl.Workbook()
# Output results
output_results_worksheet = output_workbook.create_sheet("match",  0)
if OUTPUT_WRITE_HEADER:
    output_results_worksheet["A1"].value = INVENTAIRE_PREFIX + "_serial"
    output_results_worksheet["B1"].value = PARC_PREFIXE + "_serial"
    for i in range(3,  3 + len(INVENTAIRE_OUTPUT_COLUMNS)):
        output_results_worksheet.cell(row=1,  column=i,  value=INVENTAIRE_PREFIX + "_" + INVENTAIRE_OUTPUT_COLUMNS[i - 3]["label"])
    i += 1
    for j in range(i,  i  + len(PARC_OUTPUT_COLUMNS)):
        output_results_worksheet.cell(row=1,  column=j,  value=PARC_PREFIXE + "_" + PARC_OUTPUT_COLUMNS[j-i]["label"])
if OUTPUT_WRITE_HEADER:
    decalage = 2
else:
    decalage = 1
for i in range(len(results)):
    row_cell = i + decalage
    # Write Serials : 
    output_results_worksheet["A" + str(row_cell)].value = results[i]["inv_key"]
    output_results_worksheet["B" + str(row_cell)].value = results[i]["parc_key"]
    # Write more data
    for j in range(3,  3 + len(INVENTAIRE_OUTPUT_COLUMNS)):
        output_results_worksheet.cell(row=row_cell,  column=j,  value=inventaire_sheet[INVENTAIRE_OUTPUT_COLUMNS[3-j]["col"] + str(results[i]["inv_row"])].value)
    j += 1
    for k in range(j,  j  + len(PARC_OUTPUT_COLUMNS)):
        output_results_worksheet.cell(row=row_cell,  column=k,  value=parc_sheet[PARC_OUTPUT_COLUMNS[k-j]["col"] + str(results[i]["parc_row"])].value)

# Output doublon

def doublon_output(sheet_name,  pos,  output_columns,  doublon_dict,  initial_sheet,  output_workbook):
    output_doublon_worksheet = output_workbook.create_sheet(sheet_name,  pos)

    doublon_p_struct = P_Struct()
    doublon_p_struct.write_header(output_columns,  key="label")
    for doublons in doublon_dict.values():
        for row_id in doublons:
            row = []
            for col in output_columns:
                row.append(initial_sheet[col['col'] + str(row_id)].value)
            doublon_p_struct.add_data_row(row)
    doublon_p_struct.write_in_excel_sheet(output_doublon_worksheet)

def no_match_output(sheet_name,  pos,  output_columns,  no_match_dict,  initial_sheet,  output_workbook):
    output_no_match_worksheet = output_workbook.create_sheet(sheet_name,  pos)

    no_match_p_struct = P_Struct()
    no_match_p_struct.write_header(output_columns,  key="label")
    if type(no_match_dict) is list:
        for no_match_row_id in no_match_dict:
            row = []
            for col in output_columns:
                row.append(initial_sheet[col['col'] + str(no_match_row_id)].value)
            no_match_p_struct.add_data_row(row)
    elif type(no_match_dict) is dict:
        for no_match_row_id in no_match_dict.values():
                row = []
                for col in output_columns:
                    row.append(initial_sheet[col['col'] + str(no_match_row_id)].value)
                no_match_p_struct.add_data_row(row)
    no_match_p_struct.write_in_excel_sheet(output_no_match_worksheet)

no_match_output("inv_no_match",  1,  INVENTAIRE_OUTPUT_COLUMNS,  inventaire_no_match,  inventaire_sheet,  output_workbook)    
no_match_output("parc_no_match",  2,  PARC_OUTPUT_COLUMNS,  parc_no_match,  parc_sheet, output_workbook)
doublon_output("inv_doublon",  3,   INVENTAIRE_OUTPUT_COLUMNS,  inventaire_key_dict_doublon,  inventaire_sheet,  output_workbook)
doublon_output("parc_doublon",  4,  PARC_OUTPUT_COLUMNS,  parc_key_dict_doublon,  parc_sheet,  output_workbook)
no_match_output("inv_no_serial",  5,  INVENTAIRE_OUTPUT_COLUMNS,  inventaire_no_key_dict,  inventaire_sheet,  output_workbook)
no_match_output("parc_no_serial",  6, PARC_OUTPUT_COLUMNS,  parc_no_key_dict,  parc_sheet,  output_workbook)
no_match_output("inv_short_serial",  7,  INVENTAIRE_OUTPUT_COLUMNS,  inventaire_key_dict_short,  inventaire_sheet,  output_workbook)
no_match_output("parc_short_serial",  8,  PARC_OUTPUT_COLUMNS,  parc_key_dict_short, parc_sheet,  output_workbook)


output_workbook.save(OUTPUT_FILE)
