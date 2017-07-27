import openpyxl

class P_Struct:
    def __init__(self):
        self.data = []
        self.header = None
        
    def add_empty_rows(self,  until):
        for i in range(until - len(self.data) + 1):
                self.data.append([])
                
        
    def write_header(self,  header,  mode='a'):
        if mode == 'a':
            try:
                self.data[0]
            except:
                self.data.append([])
        elif mode == 'w':
            self.data[0] = []
            
        self.header = self.data[0]
        
        for label in header:
                self.header.append(label)
                
    def add_data_row(self,  row):
        self.data.append([])
        current_row = self.data[-1]
        for cell in row:
            current_row.append(cell)
            
    def write_data_row(self,  row,  at,  mode='a'):
        if len(self.data) - 1 < at:
            # Il faut d'abord créer la rangée
            self.add_empty_rows(at)
            
        current_row = self.data[at]
        if mode == 'w':
            current_row = []
            
        for cell in row:
            current_row.append(cell)
  
    def write_in_excel_sheet(self,  output_sheet):
        for row in range(0,  len(self.data)):
            for col in range(0,  len(self.data[row])):
                output_sheet.cell(row=row + 1,  column=col + 1,  value=self.data[row][col])
                
        return output_sheet
                
test = P_Struct()
test.write_header(["test",  "test1",  "test2"])
test.write_header(["test3",  "test4",  "test5"])
test.add_data_row(["Hello",  "World",  "!"])
test.write_data_row(["1",  "2",  "3"],  5)
#test.write_header(["Hello"],  mode="w")
output_workbook = openpyxl.Workbook()
output_results_worksheet = output_workbook.create_sheet("match",  0)
output_results_worksheet = test.write_in_excel_sheet(output_results_worksheet)
output_workbook.save("test.xlsx")
